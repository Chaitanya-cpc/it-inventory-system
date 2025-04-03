import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from datetime import date, timedelta
import random

# --- Configuration ---
FILENAME = "company_asset_inventory_with_log.xlsx"
SHEET_HARDWARE = "Hardware Inventory"
SHEET_SOFTWARE = "Software & Licensing"
SHEET_AUDIT_LOG = "Audit Log" # New sheet for logging

# --- Define Headers ---
HARDWARE_HEADERS = [
    "Asset ID", "Category", "Sub-Category", "Name", "Status", "Model",
    "Serial Number / Service Tag", "Date of Purchase", "Acquisition Type",
    "Assigned To", "Location", "Notes"
]

SOFTWARE_HEADERS = [
    "License ID", "Hardware Asset ID / SN", "Type", "Product Name / Details",
    "License Key / Identifier", "Activation Date", "Expiry Date / Warranty End",
    "Quantity / Seats", "Assigned To", "Notes"
]

AUDIT_LOG_HEADERS = [ # Headers for the new log sheet
    "Timestamp", "User", "Sheet Name", "Cell Address", "Action", "Old Value", "New Value"
]


# --- Define Categories & Sample Data (Same as before) ---
HARDWARE_CATEGORIES = {
    "Laptops": ["Engineering", "Business"],
    "Desktops": ["AIO", "Mini PC", "Assembled Tower"],
    "Mouse": [None],
    "Keyboards": [None],
    "Servers": ["Rackmount", "Tower"],
    "Printers": ["Laser", "Inkjet", "Plotter", "Label"],
    "Wifi": ["Access Point", "Router", "Modem", "Dongle", "Switch"],
    "Phones": ["Desk Phone", "Mobile - Android", "Mobile - iOS"],
    "Storage Drives": ["External HDD", "SSD", "NAS Unit"],
    "Monitors": ["24 Inch", "27 Inch", "Ultrawide"],
    "UPS": ["Desktop", "Rackmount"]
}
STATUS_OPTIONS = ["New", "Okay", "Needs Repair", "Dead"]
ACQUISITION_OPTIONS = ["Purchased", "Leased", "Rented"]
SOFTWARE_TYPES = [
    "OS License", "Office Suite", "Antivirus", "Firewall",
    "Creative Suite", "Development Tool", "Network Info", "Warranty"
]

# --- Define Styles (Same as before) ---
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Blue
audit_header_fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid") # Grey for audit log
header_alignment = Alignment(horizontal="center", vertical="center")
center_alignment = Alignment(horizontal="center", vertical="center")
left_alignment = Alignment(horizontal="left", vertical="center")
thin_border = Border(left=Side(style='thin'),
                   right=Side(style='thin'),
                   top=Side(style='thin'),
                   bottom=Side(style='thin'))
status_fills = {
    "New": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), # Green
    "Okay": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"), # White
    "Needs Repair": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), # Yellow
    "Dead": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red
}
expired_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red
warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Yellow


# --- Helper Function to Apply Header Style ---
def style_header(sheet, headers, fill=header_fill): # Added fill parameter
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = fill # Use the provided fill
        cell.alignment = header_alignment
        cell.border = thin_border

# --- Helper Function to Auto-fit Columns (Same as before) ---
def auto_fit_columns(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

# --- Generate Sample Hardware Data (Same as before) ---
def generate_sample_hardware(num_rows=50):
    data = []
    used_serials = set()
    for i in range(1, num_rows + 1):
        asset_id = f"HW{i:04d}"
        category = random.choice(list(HARDWARE_CATEGORIES.keys()))
        sub_cats = HARDWARE_CATEGORIES[category]
        sub_category = random.choice(sub_cats) if sub_cats[0] is not None else ""
        name = f"{category} {i}"
        status = random.choice(STATUS_OPTIONS)
        model = f"Model-{random.choice(['X1','Pro','Air','S2','T480','Precision','Optiplex','Latitude',' LaserJet','XPS','Blade','PowerEdge','Catalyst'])} {random.randint(100,9999)}"
        serial = f"SN{random.randint(10000000, 99999999)}"
        while serial in used_serials:
             serial = f"SN{random.randint(10000000, 99999999)}"
        used_serials.add(serial)
        purchase_date = date.today() - timedelta(days=random.randint(30, 1825))
        acquisition = random.choice(ACQUISITION_OPTIONS)
        assigned_to = random.choice(["John Doe", "Jane Smith", "IT Department", "Marketing", "Engineering", "Unassigned"])
        location = random.choice(["Office 101", "HQ Floor 2", "Data Center A", "Remote", "Storage"])
        notes = "Sample Note" if random.random() > 0.8 else ""
        data.append([
            asset_id, category, sub_category, name, status, model, serial,
            purchase_date, acquisition, assigned_to, location, notes
        ])
    return data, list(used_serials)

# --- Generate Sample Software Data (Same as before) ---
def generate_sample_software(num_rows=70, hardware_serials=[]):
    data = []
    today = date.today()
    for i in range(1, num_rows + 1):
        lic_id = f"LIC{i:04d}"
        hw_serial = random.choice(hardware_serials) if hardware_serials and random.random() > 0.2 else "N/A (Site/User)"
        sw_type = random.choice(SOFTWARE_TYPES)
        product_name = f"{sw_type} Product {random.randint(1,100)}"
        identifier = f"KEY-{random.randint(10000,99999)}XXX"
        if sw_type == "Network Info":
             identifier_type = random.choice(["IP", "MAC", "IMEI", "DNS"])
             if identifier_type == "IP": identifier = f"192.168.{random.randint(1,10)}.{random.randint(2,254)}"
             elif identifier_type == "MAC": identifier = ':'.join(['%02x'%random.randint(0, 255) for _ in range(6)]).upper()
             elif identifier_type == "IMEI": identifier = f"{random.randint(100000000000000,999999999999999)}"
             else: identifier = f"dns{random.randint(1,2)}.yourcompany.com"
             product_name = identifier_type
        elif sw_type == "Warranty":
             product_name = random.choice(["Dell ProSupport", "HP CarePack", "AppleCare+", "Standard Warranty"])
        activation_date = today - timedelta(days=random.randint(10, 730))
        expiry_date = None
        if sw_type != "Network Info":
            expiry_date = activation_date + timedelta(days=random.randint(180, 1095) * random.choice([1, 3, 5]))
            if random.random() < 0.15: expiry_date = today - timedelta(days=random.randint(1, 180))
            elif random.random() < 0.25: expiry_date = today + timedelta(days=random.randint(1, 60))
        quantity = 1
        if random.random() > 0.9:
            quantity = random.choice([5, 10, 25, 50])
            hw_serial = "N/A (Volume License)"
        assigned_to = "N/A"
        if hw_serial != "N/A (Site/User)" and hw_serial != "N/A (Volume License)":
             assigned_to = f"See HW: {hw_serial}"
        notes = "Sample software note." if random.random() > 0.9 else ""
        data.append([
            lic_id, hw_serial, sw_type, product_name, identifier,
            activation_date, expiry_date, quantity, assigned_to, notes
        ])
    return data


# --- Main Script ---
if __name__ == "__main__":
    print("Generating Excel Asset Inventory with Audit Log structure...")

    # Create workbook and sheets
    wb = openpyxl.Workbook()
    hw_sheet = wb.active
    hw_sheet.title = SHEET_HARDWARE
    sw_sheet = wb.create_sheet(SHEET_SOFTWARE)
    audit_sheet = wb.create_sheet(SHEET_AUDIT_LOG) # Create the audit log sheet

    # --- Process Hardware Sheet (Same as before) ---
    print(f"Populating {SHEET_HARDWARE}...")
    style_header(hw_sheet, HARDWARE_HEADERS)
    sample_hw_data, hw_serials_list = generate_sample_hardware(num_rows=75)
    for r_idx, row_data in enumerate(sample_hw_data, 2):
        for c_idx, value in enumerate(row_data, 1):
            cell = hw_sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = left_alignment
            if isinstance(value, date):
                cell.number_format = 'YYYY-MM-DD'
                cell.alignment = center_alignment
            if HARDWARE_HEADERS[c_idx-1] == "Status":
                 cell.alignment = center_alignment
    status_col_letter = get_column_letter(HARDWARE_HEADERS.index("Status") + 1)
    status_range = f"{status_col_letter}2:{status_col_letter}{hw_sheet.max_row}"
    for status, fill in status_fills.items():
        if status != "Okay":
             hw_sheet.conditional_formatting.add(status_range,
                 CellIsRule(operator='equal', formula=[f'"{status}"'], stopIfTrue=True, fill=fill))
    hw_sheet.freeze_panes = 'A2'
    hw_sheet.auto_filter.ref = hw_sheet.dimensions
    auto_fit_columns(hw_sheet)

    # --- Process Software Sheet (Same as before) ---
    print(f"Populating {SHEET_SOFTWARE}...")
    style_header(sw_sheet, SOFTWARE_HEADERS)
    sample_sw_data = generate_sample_software(num_rows=100, hardware_serials=hw_serials_list)
    today_date = date.today()
    warning_date = today_date + timedelta(days=60)
    for r_idx, row_data in enumerate(sample_sw_data, 2):
        for c_idx, value in enumerate(row_data, 1):
            cell = sw_sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = left_alignment
            if isinstance(value, date):
                cell.number_format = 'YYYY-MM-DD'
                cell.alignment = center_alignment
            if SOFTWARE_HEADERS[c_idx-1] == "Quantity / Seats":
                 cell.alignment = center_alignment
    expiry_col_letter = get_column_letter(SOFTWARE_HEADERS.index("Expiry Date / Warranty End") + 1)
    expiry_range = f"{expiry_col_letter}2:{expiry_col_letter}{sw_sheet.max_row}"
    sw_sheet.conditional_formatting.add(expiry_range,
        CellIsRule(operator='lessThan', formula=[f'TODAY()'], stopIfTrue=True, fill=expired_fill))
    sw_sheet.conditional_formatting.add(expiry_range,
        CellIsRule(operator='lessThan', formula=[f'TODAY()+60'], stopIfTrue=True, fill=warning_fill))
    sw_sheet.freeze_panes = 'A2'
    sw_sheet.auto_filter.ref = sw_sheet.dimensions
    auto_fit_columns(sw_sheet)

    # --- Set up Audit Log Sheet ---
    print(f"Setting up {SHEET_AUDIT_LOG}...")
    style_header(audit_sheet, AUDIT_LOG_HEADERS, fill=audit_header_fill) # Use grey header
    audit_sheet.freeze_panes = 'A2'
    # Apply basic formatting to columns (can adjust widths further manually if needed)
    audit_sheet.column_dimensions['A'].width = 20 # Timestamp
    audit_sheet.column_dimensions['B'].width = 15 # User
    audit_sheet.column_dimensions['C'].width = 20 # Sheet Name
    audit_sheet.column_dimensions['D'].width = 15 # Cell Address
    audit_sheet.column_dimensions['E'].width = 15 # Action
    audit_sheet.column_dimensions['F'].width = 30 # Old Value
    audit_sheet.column_dimensions['G'].width = 30 # New Value
    # Note: Protection and hiding will be done manually in Excel

    # --- Add Instructions as a Comment (Optional but helpful) ---
    # You could add a text box or comment in the Excel file itself via openpyxl,
    # but it's often cleaner to provide instructions separately.
    # For example:
    # from openpyxl.comments import Comment
    # audit_sheet['A1'].comment = Comment("Enable Macros and add VBA code to ThisWorkbook module for change tracking. Protect this sheet and VBA project.", "Setup Instructions")


    # --- Save the Workbook ---
    try:
        # Important: Save as .xlsm to support macros!
        macro_enabled_filename = FILENAME.replace(".xlsx", ".xlsm")
        wb.save(macro_enabled_filename)
        print(f"\nSuccessfully created Excel file: '{macro_enabled_filename}'")
        print("\n--- IMPORTANT NEXT STEPS ---")
        print("1. Open the generated .xlsm file in Excel.")
        print("2. Enable Macros if prompted.")
        print("3. Add the provided VBA code to the 'ThisWorkbook' module.")
        print("4. Hide and Password Protect the 'Audit Log' sheet.")
        print("5. Password Protect the VBA Project.")
        print("-----------------------------")

    except PermissionError:
        print(f"\nError: Permission denied. Cannot save '{macro_enabled_filename}'.")
        print("Please ensure the file is not already open in Excel.")
    except Exception as e:
        print(f"\nAn error occurred while saving the file: {e}")